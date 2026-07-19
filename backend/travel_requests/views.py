from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
import openpyxl
import zipfile
import io
from .models import BaseRequest, Attachment, Passenger
from .serializers import BaseRequestSerializer, AttachmentSerializer, PassengerSerializer
from rest_framework.parsers import MultiPartParser, FormParser
from core.emails import notify_admins_new_request, notify_user_status_change


class RequestViewSet(viewsets.ModelViewSet):
    serializer_class = BaseRequestSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['request_type', 'status', 'agency']
    search_fields = ['id', 'agency__company_name']
    ordering_fields = ['created_at', 'updated_at']
    ordering = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        
        # Optimize queries with select_related and prefetch_related
        qs = BaseRequest.objects.select_related(
            'agency', 'agency__user', 'customer', 'assigned_to',
            'group_visa_details', 'individual_visa_details', 'air_ticket_details'
        ).prefetch_related(
            'attachments',
            'passengers',
            'group_visa_details__hotels',
            'group_visa_details__transports'
        )
        
        if user.role in ['SUPER_ADMIN', 'ADMIN']:
            return qs.all()
        elif user.role == 'AGENCY':
            from agencies.models import Agency
            Agency.objects.get_or_create(
                user=user,
                defaults={'company_name': 'Unknown Agency', 'contact_person': 'Owner', 'phone_number': ''}
            )
            return qs.filter(agency__user=user)
        elif user.role == 'CUSTOMER':
            return qs.filter(customer=user)
        return BaseRequest.objects.none()

    def perform_create(self, serializer):
        user = self.request.user
        if user.role == 'AGENCY':
            from agencies.models import Agency
            agency, _ = Agency.objects.get_or_create(
                user=user,
                defaults={'company_name': 'Unknown Agency', 'contact_person': 'Owner', 'phone_number': ''}
            )
            serializer.save(agency=agency)
        elif user.role == 'CUSTOMER':
            serializer.save(customer=user)
        else:
            serializer.save()
            
        # Create notification
        from notifications.models import Notification, NotificationType
        Notification.objects.create(
            user=user,
            title='Request Submitted',
            message=f'Your {serializer.instance.get_request_type_display()} request has been successfully submitted.',
            notification_type=NotificationType.REQUEST_SUBMITTED,
            related_request_id=serializer.instance.id
        )
        
        notify_admins_new_request(serializer.instance.request_type, str(serializer.instance.id))

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        from django.core.cache import cache
        from django.db.models import Count, Q
        
        cache_key = f"dashboard_stats_{request.user.id}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)

        qs = self.get_queryset()
        
        # Calculate stats in one query
        stats = qs.aggregate(
            total=Count('id'),
            pending=Count('id', filter=Q(status__in=['SUBMITTED', 'UNDER_REVIEW'])),
            processing=Count('id', filter=Q(status='PROCESSING')),
            completed=Count('id', filter=Q(status__in=['APPROVED', 'COMPLETED'])),
            rejected=Count('id', filter=Q(status='REJECTED')),
            draft=Count('id', filter=Q(status='DRAFT')),
            group_visa=Count('id', filter=Q(request_type='GROUP_VISA')),
            individual_visa=Count('id', filter=Q(request_type='INDIVIDUAL_VISA')),
            air_ticket=Count('id', filter=Q(request_type='AIR_TICKET')),
        )

        # Recent requests
        recent = qs.order_by('-created_at')[:5]
        # Avoid heavy serialization for dashboard
        recent_data = [
            {
                'id': str(r.id),
                'request_type': r.request_type,
                'status': r.status,
                'current_phase': r.current_phase,
                'created_at': r.created_at,
                'updated_at': r.updated_at,
                'agency_details': {'company_name': r.agency.company_name} if hasattr(r, 'agency') and r.agency else None,
                'customer_details': {'email': r.customer.email} if r.customer else None,
            }
            for r in recent
        ]

        data = {
            'total_requests': stats['total'] or 0,
            'pending_requests': stats['pending'] or 0,
            'processing_requests': stats['processing'] or 0,
            'completed_requests': stats['completed'] or 0,
            'rejected_requests': stats['rejected'] or 0,
            'draft_requests': stats['draft'] or 0,
            'distribution': {
                'group_visa': stats['group_visa'] or 0,
                'individual_visa': stats['individual_visa'] or 0,
                'air_ticket': stats['air_ticket'] or 0,
            },
            'recent_requests': recent_data
        }
        
        cache.set(cache_key, data, 300) # 5 mins cache
        return Response(data)

    @action(detail=True, methods=['patch'])
    def update_status(self, request, pk=None):
        if request.user.role not in ['SUPER_ADMIN', 'ADMIN']:
            return Response(status=status.HTTP_403_FORBIDDEN)
        
        obj = self.get_object()
        new_status = request.data.get('status')
        if not new_status:
            return Response({'detail': 'Status is required.'}, status=status.HTTP_400_BAD_REQUEST)
        
        obj.status = new_status
        obj.save()

        # Create notification
        from notifications.models import Notification, NotificationType
        user_to_notify = obj.customer if obj.customer else (obj.agency.user if obj.agency else None)
        if user_to_notify:
            Notification.objects.create(
                user=user_to_notify,
                title='Request Status Updated',
                message=f'Your request ({obj.id}) is now {new_status}.',
                notification_type=NotificationType.STATUS_CHANGED,
                related_request_id=obj.id
            )
            notify_user_status_change(user_to_notify.email, str(obj.id), new_status)

        return Response(BaseRequestSerializer(obj).data)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        if request.user.role != 'SUPER_ADMIN':
            return Response(status=status.HTTP_403_FORBIDDEN)
        
        queryset = self.filter_queryset(self.get_queryset())
        
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["Request ID", "Type", "Agency", "Customer", "Status", "Created At", "Assigned To"])
        
        # Passengers Sheet
        ws_passengers = wb.create_sheet(title="Passengers")
        ws_passengers.append(["Request ID", "Passenger Name", "Passport", "Nationality", "DOB", "Expiry"])
        
        # Group Visa / Hotels Sheet
        ws_hotels = wb.create_sheet(title="Hotels & Transports")
        ws_hotels.append(["Request ID", "Type", "Details", "Date/Check-in", "Check-out"])
        
        # Individual Visa
        ws_indiv = wb.create_sheet(title="Individual Visa")
        ws_indiv.append(["Request ID", "Visa Subtype", "Arrival", "Departure", "Stay Days", "IQAMA"])
        
        # Air Ticket
        ws_air = wb.create_sheet(title="Air Ticket")
        ws_air.append(["Request ID", "Origin", "Destination", "Dates", "Airline", "Notes"])
        
        for req in queryset:
            req_id_short = str(req.id).split('-')[0]
            
            # Summary
            ws_summary.append([
                str(req.id), req.get_request_type_display(),
                req.agency.company_name if hasattr(req, 'agency') and req.agency else '',
                req.customer.email if req.customer else '',
                req.get_status_display(), req.created_at.strftime('%Y-%m-%d %H:%M'),
                req.assigned_to.email if req.assigned_to else ''
            ])
            
            # Passengers
            if hasattr(req, 'passengers'):
                for p in req.passengers.all():
                    ws_passengers.append([
                        req_id_short, p.full_name, p.passport_number, p.nationality,
                        str(p.date_of_birth), str(p.passport_expiry)
                    ])
                
            # Group Visa
            if hasattr(req, 'group_visa_details') and req.group_visa_details:
                for h in req.group_visa_details.hotels.all():
                    ws_hotels.append([req_id_short, "Hotel", f"{h.hotel_name} ({h.city}) - {h.room_count} {h.room_type}", str(h.check_in), str(h.check_out)])
                for t in req.group_visa_details.transports.all():
                    ws_hotels.append([req_id_short, "Transport", f"{t.transport_type} ({t.period})", str(t.date), str(t.time)])
                    
            # Individual Visa
            if hasattr(req, 'individual_visa_details') and req.individual_visa_details:
                iv = req.individual_visa_details
                ws_indiv.append([
                    req_id_short, iv.visa_subtype, iv.arrival_flight, iv.departure_flight,
                    str(iv.stay_days), iv.iqama_id
                ])
                
            # Air Ticket
            if hasattr(req, 'air_ticket_details') and req.air_ticket_details:
                at = req.air_ticket_details
                ws_air.append([
                    req_id_short, at.origin, at.destination, f"{at.departure_date} to {at.arrival_date}",
                    at.preferred_airline, at.additional_notes
                ])
                
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Requests_Export.xlsx'
        wb.save(response)
        
        # Log to audit
        try:
            from audit.models import AuditLog
            AuditLog.objects.create(
                user=request.user,
                action="EXPORT_EXCEL",
                target_type="BaseRequest",
                details={"count": queryset.count()}
            )
        except Exception:
            pass
            
        return response

    @action(detail=True, methods=['get'])
    def download_all_documents(self, request, pk=None):
        if request.user.role not in ['SUPER_ADMIN', 'ADMIN']:
            return Response(status=status.HTTP_403_FORBIDDEN)
            
        obj = self.get_object()
        passengers = obj.passengers.exclude(passport_document='') if hasattr(obj, 'passengers') else []
        attachments = obj.attachments.all() if hasattr(obj, 'attachments') else []
        
        passengers_exists = passengers.exists() if hasattr(passengers, 'exists') else bool(passengers)
        attachments_exists = attachments.exists() if hasattr(attachments, 'exists') else bool(attachments)
        
        if not passengers_exists and not attachments_exists:
            return Response({'detail': 'No documents found for this request.'}, status=status.HTTP_404_NOT_FOUND)
            
        import requests
        zip_buffer = io.BytesIO()
        req_id_short = str(obj.id).split('-')[0]
        
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED, False) as zip_file:
            # Passengers
            if passengers_exists:
                for p in passengers:
                    if not p.passport_document: continue
                    file_url = p.passport_document.url
                    try:
                        if file_url.startswith('http'):
                            file_data = requests.get(file_url).content
                        else:
                            p.passport_document.seek(0)
                            file_data = p.passport_document.read()
                        ext = file_url.split('.')[-1]
                        if len(ext) > 4: ext = "pdf"
                        filename = f"Passengers/{p.full_name.replace(' ', '_')}_Passport.{ext}"
                        zip_file.writestr(filename, file_data)
                    except Exception as e:
                        print(f"Failed to fetch passport for {p.full_name}: {e}")
            
            # Attachments
            if attachments_exists:
                for a in attachments:
                    if not a.file: continue
                    file_url = a.file.url
                    try:
                        if file_url.startswith('http'):
                            file_data = requests.get(file_url).content
                        else:
                            a.file.seek(0)
                            file_data = a.file.read()
                        filename = f"Attachments/{a.file_name}"
                        zip_file.writestr(filename, file_data)
                    except Exception as e:
                        print(f"Failed to fetch attachment {a.file_name}: {e}")
                    
        # Log to audit
        try:
            from audit.models import AuditLog
            AuditLog.objects.create(
                user=request.user,
                action="DOWNLOAD_DOCUMENTS",
                target_type="BaseRequest",
                target_id=str(obj.id)
            )
        except Exception:
            pass
                    
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="Request_{req_id_short}_Documents.zip"'
        return response



class AttachmentViewSet(viewsets.ModelViewSet):
    serializer_class = AttachmentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ['SUPER_ADMIN', 'ADMIN']:
            return Attachment.objects.all()
        return Attachment.objects.filter(request__agency__user=user) | Attachment.objects.filter(request__customer=user)

    def perform_create(self, serializer):
        file_obj = self.request.FILES.get('file')
        if file_obj:
            serializer.save(
                uploaded_by=self.request.user,
                file_name=file_obj.name,
                file_type=file_obj.content_type,
                file_size=file_obj.size
            )
        else:
            serializer.save(uploaded_by=self.request.user)

    def perform_update(self, serializer):
        file_obj = self.request.FILES.get('file')
        if file_obj:
            serializer.save(
                file_name=file_obj.name,
                file_type=file_obj.content_type,
                file_size=file_obj.size
            )
        else:
            serializer.save()

class PassengerViewSet(viewsets.ModelViewSet):
    serializer_class = PassengerSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        user = self.request.user
        if user.role in ['SUPER_ADMIN', 'ADMIN']:
            return Passenger.objects.all()
        return Passenger.objects.filter(request__agency__user=user) | Passenger.objects.filter(request__customer=user)

    @action(detail=True, methods=['patch'])
    def upload(self, request, pk=None):
        passenger = self.get_object()
        file_obj = request.FILES.get('passport_document')
        if not file_obj:
            return Response({'detail': 'No document provided.'}, status=status.HTTP_400_BAD_REQUEST)
        
        passenger.passport_document = file_obj
        passenger.save()
        return Response(PassengerSerializer(passenger).data)
